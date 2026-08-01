"""General batch-weight calculator for solid-state ceramic synthesis.

Give it any target composition (as a formula string) and a batch mass, and it
works out how many grams of each raw material to weigh out.  Every run is
appended to the log workbook - nothing already in the file is overwritten.

Quick use - edit the CONFIG block below and press run, or from a terminal:

    python stoichiometry.py "Sr0.38La0.12Ba0.5Ti0.12Nb1.88O6" 30
    python stoichiometry.py "Ba0.5Sr0.5TiO3" 25 --use Ti=TiO2 --purity Nb2O5=0.998
    python stoichiometry.py "BaTiO3" 10 --note "calcine 1200 C, 4 h"
"""

from __future__ import annotations

import argparse
import re
import sys
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# CONFIG - the only part you normally need to touch
# ---------------------------------------------------------------------------

COMPOSITION = "Sr0.38 La0.12 Ba0.5 Ti0.12 Nb1.88 O6"
TARGET_MASS_G = 30.0

# Override the default raw material for a cation, e.g. {"Ti": "TiO2", "Ba": "BaCO3"}.
PRECURSOR_CHOICES: dict[str, str] = {}

# Purity as a fraction, keyed by precursor formula, e.g. {"Nb2O5": 0.9985}.
# Reagents are taken as 100% pure unless you say otherwise, so batches stay
# comparable run to run.  A purity common to every reagent cancels out of the
# cation ratio anyway - only differences between reagents shift stoichiometry.
PURITIES: dict[str, float] = {}
DEFAULT_PURITY = 1.0

LOG_FILE = "stoichiometry_log.xlsx"
NOTES = ""

# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------

# Standard atomic weights (IUPAC), g/mol.
ATOMIC_MASS: dict[str, float] = {
    "H": 1.008,
    "He": 4.0026,
    "Li": 6.94,
    "Be": 9.0122,
    "B": 10.81,
    "C": 12.011,
    "N": 14.007,
    "O": 15.999,
    "F": 18.998,
    "Ne": 20.180,
    "Na": 22.990,
    "Mg": 24.305,
    "Al": 26.982,
    "Si": 28.085,
    "P": 30.974,
    "S": 32.06,
    "Cl": 35.45,
    "Ar": 39.95,
    "K": 39.098,
    "Ca": 40.078,
    "Sc": 44.956,
    "Ti": 47.867,
    "V": 50.942,
    "Cr": 51.996,
    "Mn": 54.938,
    "Fe": 55.845,
    "Co": 58.933,
    "Ni": 58.693,
    "Cu": 63.546,
    "Zn": 65.38,
    "Ga": 69.723,
    "Ge": 72.630,
    "As": 74.922,
    "Se": 78.971,
    "Br": 79.904,
    "Kr": 83.798,
    "Rb": 85.468,
    "Sr": 87.62,
    "Y": 88.906,
    "Zr": 91.224,
    "Nb": 92.906,
    "Mo": 95.95,
    "Tc": 98.0,
    "Ru": 101.07,
    "Rh": 102.91,
    "Pd": 106.42,
    "Ag": 107.87,
    "Cd": 112.41,
    "In": 114.82,
    "Sn": 118.71,
    "Sb": 121.76,
    "Te": 127.60,
    "I": 126.90,
    "Xe": 131.29,
    "Cs": 132.91,
    "Ba": 137.327,
    "La": 138.905,
    "Ce": 140.116,
    "Pr": 140.908,
    "Nd": 144.242,
    "Pm": 145.0,
    "Sm": 150.36,
    "Eu": 151.964,
    "Gd": 157.25,
    "Tb": 158.925,
    "Dy": 162.500,
    "Ho": 164.930,
    "Er": 167.259,
    "Tm": 168.934,
    "Yb": 173.045,
    "Lu": 174.967,
    "Hf": 178.486,
    "Ta": 180.948,
    "W": 183.84,
    "Re": 186.207,
    "Os": 190.23,
    "Ir": 192.217,
    "Pt": 195.084,
    "Au": 196.967,
    "Hg": 200.592,
    "Tl": 204.38,
    "Pb": 207.2,
    "Bi": 208.980,
    "Po": 209.0,
    "Th": 232.038,
    "Pa": 231.036,
    "U": 238.029,
}

# Default raw material for each cation.  Anything here can be overridden per run.
DEFAULT_PRECURSOR: dict[str, str] = {
    "Li": "Li2CO3",
    "Na": "Na2CO3",
    "K": "K2CO3",
    "Rb": "Rb2CO3",
    "Cs": "Cs2CO3",
    "Mg": "MgO",
    "Ca": "CaCO3",
    "Sr": "SrCO3",
    "Ba": "BaCO3",
    "Sc": "Sc2O3",
    "Y": "Y2O3",
    "Ti": "TiO2",
    "Zr": "ZrO2",
    "Hf": "HfO2",
    "V": "V2O5",
    "Nb": "Nb2O5",
    "Ta": "Ta2O5",
    "Cr": "Cr2O3",
    "Mo": "MoO3",
    "W": "WO3",
    "Mn": "MnO2",
    "Fe": "Fe2O3",
    "Co": "Co3O4",
    "Ni": "NiO",
    "Cu": "CuO",
    "Zn": "ZnO",
    "Al": "Al2O3",
    "Ga": "Ga2O3",
    "In": "In2O3",
    "B": "H3BO3",
    "Si": "SiO2",
    "Ge": "GeO2",
    "Sn": "SnO2",
    "Pb": "PbO",
    "P": "NH4H2PO4",
    "Sb": "Sb2O3",
    "Bi": "Bi2O3",
    "Te": "TeO2",
    "La": "La2O3",
    "Ce": "CeO2",
    "Pr": "Pr6O11",
    "Nd": "Nd2O3",
    "Sm": "Sm2O3",
    "Eu": "Eu2O3",
    "Gd": "Gd2O3",
    "Tb": "Tb4O7",
    "Dy": "Dy2O3",
    "Ho": "Ho2O3",
    "Er": "Er2O3",
    "Tm": "Tm2O3",
    "Yb": "Yb2O3",
    "Lu": "Lu2O3",
    "Th": "ThO2",
    "U": "U3O8",
}

# Friendly names for the printout / log.  Unlisted formulae print as themselves.
PRECURSOR_NAME: dict[str, str] = {
    "Li2CO3": "Lithium carbonate",
    "Na2CO3": "Sodium carbonate",
    "K2CO3": "Potassium carbonate",
    "CaCO3": "Calcium carbonate",
    "SrCO3": "Strontium carbonate",
    "BaCO3": "Barium carbonate",
    "MgO": "Magnesium oxide",
    "Sc2O3": "Scandium oxide",
    "Y2O3": "Yttrium oxide",
    "TiO2": "Titanium dioxide",
    "ZrO2": "Zirconium dioxide",
    "HfO2": "Hafnium dioxide",
    "V2O5": "Vanadium pentoxide",
    "Nb2O5": "Niobium pentoxide",
    "Ta2O5": "Tantalum pentoxide",
    "Cr2O3": "Chromium(III) oxide",
    "MoO3": "Molybdenum trioxide",
    "WO3": "Tungsten trioxide",
    "MnO2": "Manganese dioxide",
    "Fe2O3": "Iron(III) oxide",
    "Co3O4": "Cobalt(II,III) oxide",
    "NiO": "Nickel(II) oxide",
    "CuO": "Copper(II) oxide",
    "ZnO": "Zinc oxide",
    "Al2O3": "Aluminium oxide",
    "Ga2O3": "Gallium oxide",
    "In2O3": "Indium oxide",
    "H3BO3": "Boric acid",
    "SiO2": "Silicon dioxide",
    "GeO2": "Germanium dioxide",
    "SnO2": "Tin(IV) oxide",
    "PbO": "Lead(II) oxide",
    "NH4H2PO4": "Ammonium dihydrogen phosphate",
    "Sb2O3": "Antimony(III) oxide",
    "Bi2O3": "Bismuth(III) oxide",
    "TeO2": "Tellurium dioxide",
    "La2O3": "Lanthanum oxide",
    "CeO2": "Cerium(IV) oxide",
    "Pr6O11": "Praseodymium oxide",
    "Nd2O3": "Neodymium oxide",
    "Sm2O3": "Samarium oxide",
    "Eu2O3": "Europium oxide",
    "Gd2O3": "Gadolinium oxide",
    "Tb4O7": "Terbium oxide",
    "Dy2O3": "Dysprosium oxide",
    "Ho2O3": "Holmium oxide",
    "Er2O3": "Erbium oxide",
    "Tm2O3": "Thulium oxide",
    "Yb2O3": "Ytterbium oxide",
    "Lu2O3": "Lutetium oxide",
    "ThO2": "Thorium dioxide",
    "U3O8": "Triuranium octoxide",
}

# Elements supplied by the precursors themselves / lost as gas on firing, so they
# never need a raw material of their own.
VOLATILE_OR_LATTICE = {"O", "C", "H", "N"}

# ---------------------------------------------------------------------------
# Formula handling
# ---------------------------------------------------------------------------

_TOKEN = re.compile(r"([A-Z][a-z]?)(\d*\.?\d*)|(\()|(\)(\d*\.?\d*))")


def parse_formula(formula: str) -> dict[str, float]:
    """Turn a formula string into {element: atoms per formula unit}.

    Handles decimals, nested brackets and hydrates:
    ``Ba0.5Sr0.5TiO3``, ``La(OH)3``, ``Y2O3 * 3 H2O``, ``CuSO4 · 5 H2O``.

    A ``.`` between two digits is always read as a decimal point, since decimal
    subscripts are the common case here - write hydrates with ``*`` or ``·``
    (``CuSO4*5H2O``), not ``CuSO4.5H2O``, so the coefficient is unambiguous.
    """
    text = formula.replace(" ", "").replace("[", "(").replace("]", ")")

    # Hydrate notation: middot/asterisk always separates; a bare dot only when
    # it is not sitting between two digits (i.e. it is not a decimal point).
    parts = re.split(r"[·∙*]|(?<!\d)\.|\.(?!\d)", text)
    parts = [p for p in parts if p]

    total: dict[str, float] = {}
    for part in parts:
        lead = re.match(r"^(\d+\.?\d*)(?=[A-Z(])", part)
        multiplier = float(lead.group(1)) if lead else 1.0
        if lead:
            part = part[lead.end() :]
        for element, count in _parse_group(part, formula).items():
            total[element] = total.get(element, 0.0) + count * multiplier

    if not total:
        raise ValueError(f"No elements found in formula {formula!r}.")
    unknown = sorted(e for e in total if e not in ATOMIC_MASS)
    if unknown:
        raise ValueError(f"Unknown element(s) in {formula!r}: {', '.join(unknown)}")
    return total


def _parse_group(text: str, original: str) -> dict[str, float]:
    """Recursive-descent walk over one bracketed formula fragment."""
    stack: list[dict[str, float]] = [{}]
    pos = 0
    while pos < len(text):
        match = _TOKEN.match(text, pos)
        if match is None or match.end() == pos:
            raise ValueError(f"Could not parse {original!r} at {text[pos:]!r}.")
        pos = match.end()

        if match.group(3):  # opening bracket
            stack.append({})
        elif match.group(4):  # closing bracket + count
            count = float(match.group(5)) if match.group(5) else 1.0
            if len(stack) == 1:
                raise ValueError(f"Unbalanced brackets in {original!r}.")
            group = stack.pop()
            for element, amount in group.items():
                stack[-1][element] = stack[-1].get(element, 0.0) + amount * count
        else:  # element symbol + count
            element = match.group(1)
            count = float(match.group(2)) if match.group(2) else 1.0
            stack[-1][element] = stack[-1].get(element, 0.0) + count

    if len(stack) != 1:
        raise ValueError(f"Unbalanced brackets in {original!r}.")
    return stack[0]


def molar_mass(formula: str | dict[str, float]) -> float:
    """Molar mass in g/mol of a formula string or an already-parsed composition."""
    atoms = parse_formula(formula) if isinstance(formula, str) else formula
    return sum(count * ATOMIC_MASS[element] for element, count in atoms.items())


def format_composition(atoms: dict[str, float]) -> str:
    """Tidy one-line rendering of a parsed composition."""

    def fmt(value: float) -> str:
        return "" if value == 1 else f"{value:g}"

    return "".join(f"{element}{fmt(count)}" for element, count in atoms.items())


# ---------------------------------------------------------------------------
# The calculation
# ---------------------------------------------------------------------------


@dataclass
class Reagent:
    """One raw material to weigh out."""

    name: str
    formula: str
    cations: list[str]
    molar_mass: float
    purity: float
    moles: float
    mass_g: float


@dataclass
class Batch:
    """Everything one run of the calculation produced."""

    composition: str
    atoms: dict[str, float]
    formula_weight: float
    target_mass_g: float
    moles: float
    reagents: list[Reagent]
    notes: str = ""
    warnings: list[str] = field(default_factory=list)

    @property
    def total_reagent_mass_g(self) -> float:
        return sum(r.mass_g for r in self.reagents)


def calculate_batch(
    composition: str,
    target_mass_g: float,
    precursor_choices: dict[str, str] | None = None,
    purities: dict[str, float] | None = None,
    default_purity: float = DEFAULT_PURITY,
    notes: str = "",
) -> Batch:
    """Work out the raw materials needed for `target_mass_g` of `composition`."""
    if target_mass_g <= 0:
        raise ValueError("Target batch mass must be greater than zero.")

    precursor_choices = {**(precursor_choices or {})}
    purities = purities or {}

    atoms = parse_formula(composition)
    formula_weight = molar_mass(atoms)
    moles_product = target_mass_g / formula_weight

    warnings: list[str] = []

    # 1. Pick a raw material for every element that needs one.
    chosen: dict[str, str] = {}
    for element in atoms:
        if element in precursor_choices:
            chosen[element] = precursor_choices[element]
        elif element in DEFAULT_PRECURSOR:
            chosen[element] = DEFAULT_PRECURSOR[element]
        elif element in VOLATILE_OR_LATTICE:
            continue  # supplied by the oxides/carbonates or by the furnace air
        else:
            raise ValueError(
                f"No default raw material for {element}. Pass one explicitly, "
                f'e.g. PRECURSOR_CHOICES = {{"{element}": "{element}O2"}} '
                f"or --use {element}={element}O2."
            )

    # 2. Group cations that share a raw material (e.g. BaTiO3 supplies Ba and Ti).
    grouped: dict[str, list[str]] = {}
    for element, formula in chosen.items():
        grouped.setdefault(formula, []).append(element)

    reagents: list[Reagent] = []
    for formula, elements in grouped.items():
        reagent_atoms = parse_formula(formula)
        reagent_mass = molar_mass(reagent_atoms)
        purity = purities.get(formula, default_purity)
        if not 0 < purity <= 1:
            raise ValueError(f"Purity for {formula} must be between 0 and 1, got {purity}.")

        demands = {}
        for element in elements:
            per_reagent = reagent_atoms.get(element)
            if not per_reagent:
                raise ValueError(f"{formula} contains no {element}.")
            demands[element] = moles_product * atoms[element] / per_reagent

        moles_reagent = max(demands.values())
        if max(demands.values()) - min(demands.values()) > 1e-9 * moles_reagent:
            detail = ", ".join(f"{e}: {v:.6g} mol" for e, v in demands.items())
            warnings.append(
                f"{formula} supplies {'/'.join(elements)} in a ratio that does not match "
                f"the target ({detail}). Using the largest and leaving the rest in excess - "
                f"add a second raw material for the shortfall."
            )

        reagents.append(
            Reagent(
                name=PRECURSOR_NAME.get(formula, formula),
                formula=formula,
                cations=sorted(elements),
                molar_mass=reagent_mass,
                purity=purity,
                moles=moles_reagent,
                mass_g=moles_reagent * reagent_mass / purity,
            )
        )

    # 3. Balance check: what the reagents actually deliver vs what the target wants.
    #    Catches missing elements, and raw materials that bring in a second cation
    #    on top of the one they were picked for (e.g. BaTiO3 alongside TiO2).
    delivered: dict[str, float] = {}
    for reagent in reagents:
        for element, count in parse_formula(reagent.formula).items():
            delivered[element] = delivered.get(element, 0.0) + reagent.moles * count

    for element in sorted(set(atoms) | set(delivered)):
        if element in VOLATILE_OR_LATTICE:
            continue  # O/C/H/N come and go with the furnace atmosphere
        wanted = moles_product * atoms.get(element, 0.0)
        got = delivered.get(element, 0.0)
        if abs(got - wanted) <= 1e-9 * max(wanted, got, 1e-12):
            continue
        if wanted == 0:
            warnings.append(f"{element} is not in the target but the raw materials add it.")
        elif got == 0:
            warnings.append(f"Nothing in the batch supplies {element}.")
        else:
            warnings.append(
                f"{element} is {'over' if got > wanted else 'under'}-supplied by "
                f"{abs(got - wanted) / wanted:.1%} ({got:.6g} mol delivered vs "
                f"{wanted:.6g} mol wanted) - adjust the raw materials for {element}."
            )

    reagents.sort(key=lambda r: r.name.lower())

    return Batch(
        composition=composition.strip(),
        atoms=atoms,
        formula_weight=formula_weight,
        target_mass_g=target_mass_g,
        moles=moles_product,
        reagents=reagents,
        notes=notes,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------


def print_report(batch: Batch) -> None:
    """Print the composition, the batch mass in grams, then the raw materials."""
    width = 78
    print()
    print("=" * width)
    print(f"  {format_composition(batch.atoms)}")
    print("=" * width)
    print(f"  Composition entered : {batch.composition}")
    print(f"  Formula weight      : {batch.formula_weight:.3f} g/mol")
    print(f"  Target batch mass   : {batch.target_mass_g:.4f} g")
    print(f"  Moles of product    : {batch.moles:.6f} mol")
    if batch.notes:
        print(f"  Notes               : {batch.notes}")

    print()
    print("  RAW MATERIALS REQUIRED")
    print("  " + "-" * (width - 2))
    print(f"  {'Raw material':<28}{'Formula':<12}{'M (g/mol)':>11}{'Purity':>9}{'Mass (g)':>12}")
    print("  " + "-" * (width - 2))
    for reagent in batch.reagents:
        print(
            f"  {reagent.name[:27]:<28}{reagent.formula:<12}{reagent.molar_mass:>11.3f}"
            f"{reagent.purity:>9.4f}{reagent.mass_g:>12.4f}"
        )
    print("  " + "-" * (width - 2))
    print(f"  {'Total weighed mass':<60}{batch.total_reagent_mass_g:>12.4f}")
    loss = batch.total_reagent_mass_g - batch.target_mass_g
    print(
        f"  {'Mass lost on firing (CO2, H2O, ...)':<60}{loss:>12.4f}"
        if loss > 1e-9
        else f"  {'Mass change on firing':<60}{loss:>12.4f}"
    )

    for warning in batch.warnings:
        print(f"\n  ! {warning}")
    print()


# ---------------------------------------------------------------------------
# Logging - always appends, never overwrites what is already in the workbook
# ---------------------------------------------------------------------------

BATCH_HEADERS = [
    "Batch ID",
    "Timestamp",
    "Composition",
    "Target Mass (g)",
    "Formula Weight (g/mol)",
    "Moles of Product",
    "Total Weighed Mass (g)",
    "Number of Raw Materials",
    "Notes",
]

REAGENT_HEADERS = [
    "Batch ID",
    "Timestamp",
    "Composition",
    "Raw Material",
    "Formula",
    "Supplies",
    "Molar Mass (g/mol)",
    "Purity",
    "Moles",
    "Mass (g)",
]


def _get_sheet(workbook: Workbook, title: str, headers: list[str]):
    """Fetch a sheet by name, creating it with a header row if it is not there."""
    if title in workbook.sheetnames:
        return workbook[title]
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(
            12, len(header) + 2
        )
    sheet.freeze_panes = "A2"
    return sheet


def _next_batch_id(sheet) -> int:
    ids = [row[0] for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)]
    numeric = [int(value) for value in ids if isinstance(value, (int, float))]
    return max(numeric, default=0) + 1


def append_to_log(batch: Batch, filename: str = LOG_FILE) -> tuple[Path, int]:
    """Append this batch to the workbook. Existing sheets and rows are untouched."""
    path = Path(filename)

    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        for name in list(workbook.sheetnames):  # drop the empty default sheet
            workbook.remove(workbook[name])

    batches = _get_sheet(workbook, "Batches", BATCH_HEADERS)
    materials = _get_sheet(workbook, "Raw Materials", REAGENT_HEADERS)

    batch_id = _next_batch_id(batches)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    batches.append(
        [
            batch_id,
            timestamp,
            batch.composition,
            round(batch.target_mass_g, 4),
            round(batch.formula_weight, 4),
            round(batch.moles, 8),
            round(batch.total_reagent_mass_g, 4),
            len(batch.reagents),
            batch.notes,
        ]
    )

    for reagent in batch.reagents:
        materials.append(
            [
                batch_id,
                timestamp,
                batch.composition,
                reagent.name,
                reagent.formula,
                ", ".join(reagent.cations),
                round(reagent.molar_mass, 4),
                reagent.purity,
                round(reagent.moles, 8),
                round(reagent.mass_g, 4),
            ]
        )

    try:
        workbook.save(path)
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}")
        workbook.save(fallback)
        raise PermissionError(
            f"{path} is locked (is it open in Excel?). Saved to {fallback} instead."
        ) from None

    return path, batch_id


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def _parse_pairs(values: list[str] | None, name: str, cast: Callable[[str], object] = str) -> dict:
    pairs = {}
    for item in values or []:
        if "=" not in item:
            raise SystemExit(f"--{name} expects KEY=VALUE, got {item!r}.")
        key, _, value = item.partition("=")
        pairs[key.strip()] = cast(value.strip())
    return pairs


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Batch-weight calculator for any ceramic composition.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "composition", nargs="?", default=COMPOSITION, help="target formula, e.g. Ba0.5Sr0.5TiO3"
    )
    parser.add_argument(
        "mass", nargs="?", type=float, default=TARGET_MASS_G, help="batch mass in grams"
    )
    parser.add_argument(
        "--use",
        action="append",
        metavar="EL=FORMULA",
        help="raw material for an element, e.g. --use Ti=TiO2",
    )
    parser.add_argument(
        "--purity",
        action="append",
        metavar="FORMULA=FRACTION",
        help="purity of a raw material, e.g. --purity Nb2O5=0.998",
    )
    parser.add_argument(
        "--default-purity",
        type=float,
        default=DEFAULT_PURITY,
        help=f"purity used when not given (default {DEFAULT_PURITY})",
    )
    parser.add_argument("--note", default=NOTES, help="free text stored with the batch")
    parser.add_argument("--log", default=LOG_FILE, help=f"log workbook (default {LOG_FILE})")
    parser.add_argument("--no-log", action="store_true", help="print only, do not save")
    args = parser.parse_args(argv)

    choices = {**PRECURSOR_CHOICES, **_parse_pairs(args.use, "use")}
    purities = {**PURITIES, **_parse_pairs(args.purity, "purity", float)}

    try:
        batch = calculate_batch(
            composition=args.composition,
            target_mass_g=args.mass,
            precursor_choices=choices,
            purities=purities,
            default_purity=args.default_purity,
            notes=args.note,
        )
    except ValueError as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    print_report(batch)

    if not args.no_log:
        try:
            path, batch_id = append_to_log(batch, args.log)
        except PermissionError as error:
            print(f"  ! {error}\n")
            return 1
        print(f"  Appended as batch {batch_id} to {path} (sheets 'Batches' and 'Raw Materials').\n")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
