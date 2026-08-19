"""Append-only Excel logging of calculated batches."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .calculate import Batch

DEFAULT_LOG_FILE = "stoichiometry_log.xlsx"

BATCH_HEADERS = [
    "Batch ID",
    "Timestamp",
    "Composition",
    "Target Mass (g)",
    "Formula Weight (g/mol)",
    "Moles of Product",
    "Total Weighed Mass (g)",
    "Number of Raw Materials",
    "Notes",
]

REAGENT_HEADERS = [
    "Batch ID",
    "Timestamp",
    "Composition",
    "Raw Material",
    "Formula",
    "Supplies",
    "Molar Mass (g/mol)",
    "Purity",
    "Moles",
    "Mass (g)",
]


def _get_sheet(workbook: Workbook, title: str, headers: list[str]):
    """Fetch a sheet by name, creating it with a header row if it is not there."""
    if title in workbook.sheetnames:
        return workbook[title]
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(
            12, len(header) + 2
        )
    sheet.freeze_panes = "A2"
    return sheet


def _next_batch_id(sheet) -> int:
    ids = [row[0] for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)]
    numeric = [int(value) for value in ids if isinstance(value, (int, float))]
    return max(numeric, default=0) + 1


def append_to_log(batch: Batch, filename: str = DEFAULT_LOG_FILE) -> tuple[Path, int]:
    """Append this batch to the workbook. Existing sheets and rows are untouched."""
    path = Path(filename)

    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        for name in list(workbook.sheetnames):  # drop the empty default sheet
            workbook.remove(workbook[name])

    batches = _get_sheet(workbook, "Batches", BATCH_HEADERS)
    materials = _get_sheet(workbook, "Raw Materials", REAGENT_HEADERS)

    batch_id = _next_batch_id(batches)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    batches.append(
        [
            batch_id,
            timestamp,
            batch.composition,
            round(batch.target_mass_g, 4),
            round(batch.formula_weight, 4),
            round(batch.moles, 8),
            round(batch.total_reagent_mass_g, 4),
            len(batch.reagents),
            batch.notes,
        ]
    )

    for reagent in batch.reagents:
        materials.append(
            [
                batch_id,
                timestamp,
                batch.composition,
                reagent.name,
                reagent.formula,
                ", ".join(reagent.cations),
                round(reagent.molar_mass, 4),
                reagent.purity,
                round(reagent.moles, 8),
                round(reagent.mass_g, 4),
            ]
        )

    try:
        workbook.save(path)
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}")
        workbook.save(fallback)
        raise PermissionError(
            f"{path} is locked (is it open in Excel?). Saved to {fallback} instead."
        ) from None

    return path, batch_id
