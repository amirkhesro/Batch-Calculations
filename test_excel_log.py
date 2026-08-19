"""Logging tests: append-only behaviour and sheet structure."""

from openpyxl import load_workbook

from batch_calculations import append_to_log, calculate_batch


def test_log_created_and_appended(tmp_path):
    log = tmp_path / "log.xlsx"
    batch = calculate_batch("BaTiO3", 10.0)

    path, first_id = append_to_log(batch, str(log))
    assert path == log and first_id == 1

    _, second_id = append_to_log(batch, str(log))
    assert second_id == 2

    workbook = load_workbook(log)
    assert set(workbook.sheetnames) == {"Batches", "Raw Materials"}
    batches = workbook["Batches"]
    assert batches.max_row == 3  # header + two runs
    materials = workbook["Raw Materials"]
    assert materials.max_row == 1 + 2 * len(batch.reagents)
